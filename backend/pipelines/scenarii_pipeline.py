import os
import re
import tempfile
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from docx import Document
from pydantic import BaseModel

import llm_client


def _extract_structure(docx_path: Path) -> dict[str, list[dict]]:
    """Extrage ierarhia de heading-uri + textul de corp, grupate pe secțiunea H1."""
    doc = Document(str(docx_path))
    modules: dict[str, list[dict]] = defaultdict(list)
    current_h1 = "General"
    current_cap: dict | None = None
    current_sub: dict | None = None

    for p in doc.paragraphs:
        style = p.style.name if p.style else ""
        text = p.text.strip()
        if not text:
            continue

        if re.match(r"Heading 1", style, re.I):
            current_h1 = text
            current_cap = None
            current_sub = None
        elif re.match(r"Heading 2", style, re.I):
            current_cap = {"titlu": text, "text": [], "subcapitole": []}
            current_sub = None
            modules[current_h1].append(current_cap)
        elif re.match(r"Heading [3-9]", style, re.I):
            current_sub = {"titlu": text, "text": []}
            if current_cap is not None:
                current_cap["subcapitole"].append(current_sub)
            else:
                current_cap = {"titlu": text, "text": [], "subcapitole": []}
                current_sub = None
                modules[current_h1].append(current_cap)
        else:
            # Paragraf de corp — se ataseaza celui mai specific heading curent
            if current_sub is not None:
                current_sub["text"].append(text)
            elif current_cap is not None:
                current_cap["text"].append(text)
            else:
                # Text direct sub H1, fara H2/H3 — capitol implicit cu titlul modulului
                current_cap = {"titlu": current_h1, "text": [text], "subcapitole": []}
                modules[current_h1].append(current_cap)

    return dict(modules)


CHUNK_INPUT_TOKENS = 15_000  # input maxim per apel AI (~33k caractere)


def _cap_tokens(cap: dict) -> int:
    chars = len(cap["titlu"]) + sum(len(t) for t in cap["text"])
    for sub in cap["subcapitole"]:
        chars += len(sub["titlu"]) + sum(len(t) for t in sub["text"])
    return llm_client.estimate_tokens(" " * chars)


def _split_huge_sub(sub: dict) -> list[dict]:
    """Sparge un subcapitol urias pe paragrafe, in parti sub limita."""
    parts: list[dict] = []
    current: list[str] = []
    current_tokens = 0
    for para in sub["text"]:
        p_tokens = llm_client.estimate_tokens(para)
        if current and current_tokens + p_tokens > CHUNK_INPUT_TOKENS:
            titlu = sub["titlu"] if not parts else f"{sub['titlu']} (continuare)"
            parts.append({"titlu": titlu, "text": current})
            current, current_tokens = [], 0
        current.append(para)
        current_tokens += p_tokens
    titlu = sub["titlu"] if not parts else f"{sub['titlu']} (continuare)"
    parts.append({"titlu": titlu, "text": current})
    return parts


def _split_huge_cap(cap: dict) -> list[dict]:
    """Sparge un capitol urias la granite de subcapitol (si mai jos, pe paragrafe)."""
    # Text propriu urias -> il spargem direct pe paragrafe (capitol fara/pe langa subcapitole)
    own_tokens = llm_client.estimate_tokens(
        " " * (len(cap["titlu"]) + sum(len(t) for t in cap["text"]))
    )
    if own_tokens > CHUNK_INPUT_TOKENS:
        text_parts = _split_huge_sub({"titlu": cap["titlu"], "text": cap["text"]})
        pieces = [
            {"titlu": p["titlu"], "text": p["text"], "subcapitole": []}
            for p in text_parts
        ]
        if cap["subcapitole"]:
            pieces.extend(
                _split_huge_cap({
                    "titlu": f"{cap['titlu']} (continuare)",
                    "text": [],
                    "subcapitole": cap["subcapitole"],
                })
            )
        return pieces

    pieces: list[dict] = []
    current = {"titlu": cap["titlu"], "text": list(cap["text"]), "subcapitole": []}
    current_tokens = _cap_tokens(current)

    subs: list[dict] = []
    for sub in cap["subcapitole"]:
        sub_tokens = llm_client.estimate_tokens(
            " " * (len(sub["titlu"]) + sum(len(t) for t in sub["text"]))
        )
        if sub_tokens > CHUNK_INPUT_TOKENS:
            subs.extend(_split_huge_sub(sub))
        else:
            subs.append(sub)

    for sub in subs:
        sub_tokens = llm_client.estimate_tokens(
            " " * (len(sub["titlu"]) + sum(len(t) for t in sub["text"]))
        )
        if current["subcapitole"] and current_tokens + sub_tokens > CHUNK_INPUT_TOKENS:
            pieces.append(current)
            current = {"titlu": f"{cap['titlu']} (continuare)", "text": [], "subcapitole": []}
            current_tokens = 0
        current["subcapitole"].append(sub)
        current_tokens += sub_tokens
    pieces.append(current)
    return pieces


def _pack_chunks(structure: dict[str, list[dict]]) -> list[dict]:
    """Grupeaza capitolele fiecarui modul in bucati <= CHUNK_INPUT_TOKENS.

    Granite naturale: capitol (H2); capitol urias -> subcapitole; subcapitol
    urias -> paragrafe. Invariant: niciun paragraf pierdut sau duplicat.
    """
    chunks: list[dict] = []
    for modul, capitole in structure.items():
        pieces: list[dict] = []
        for cap in capitole:
            if _cap_tokens(cap) > CHUNK_INPUT_TOKENS:
                pieces.extend(_split_huge_cap(cap))
            else:
                pieces.append(cap)

        current: list[dict] = []
        current_tokens = 0
        for cap in pieces:
            cap_tokens = _cap_tokens(cap)
            if current and current_tokens + cap_tokens > CHUNK_INPUT_TOKENS:
                chunks.append({"modul": modul, "capitole": current})
                current, current_tokens = [], 0
            current.append(cap)
            current_tokens += cap_tokens
        if current:
            chunks.append({"modul": modul, "capitole": current})
    return chunks


CALL_OVERHEAD_TOKENS = 1200   # system prompt + structura JSON ceruta
OUT_TOKENS_PER_CALL = 6000  # buget de raspuns per apel AI


class _Scenariu(BaseModel):
    capitol: str = ""
    subcapitol: str = ""
    titlu_scenariu: str
    obiectiv: str = ""
    preconditii: str = ""
    pasi: str = ""
    rezultat_asteptat: str = ""
    tip_test: str = "Funcțional - Pozitiv"
    prioritate: str = "High"
    dependente: str = "—"
    observatii: str = ""


_SYSTEM_PROMPT = """Ești inginer QA senior pentru aplicații ERP (Charisma). Primești un fragment \
de specificație funcțională în limba română, structurat pe capitole și subcapitole.
Generează scenarii de testare concrete, STRICT pe baza textului primit — nu inventa funcționalități.
Pentru fiecare capitol/subcapitol: cazul pozitiv principal și, unde textul menționează validări, \
restricții sau reguli, câte un caz negativ.
Scrie în română, cu diacritice. Pașii sunt numerotați, precondițiile cu bullet •.
Răspunde DOAR cu JSON valid, fără alt text:
{"scenarii": [{"capitol": "...", "subcapitol": "...", "titlu_scenariu": "...", "obiectiv": "...", \
"preconditii": "• ...", "pasi": "1. ...\\n2. ...", "rezultat_asteptat": "...", \
"tip_test": "Funcțional - Pozitiv" sau "Funcțional - Negativ", \
"prioritate": "Critical"|"High"|"Medium"|"Low", "dependente": "...", "observatii": "..."}]}"""


def _chunk_prompt(modul: str, capitole: list[dict], part: int, total: int) -> str:
    header = f"MODUL: {modul}"
    if total > 1:
        header += f" (partea {part} din {total})"
    lines = [header, ""]
    for cap in capitole:
        lines.append(f"CAPITOL: {cap['titlu']}")
        lines.extend(cap["text"])
        for sub in cap["subcapitole"]:
            lines.append(f"SUBCAPITOL: {sub['titlu']}")
            lines.extend(sub["text"])
        lines.append("")
    return "\n".join(lines)


def _module_stubs(modul: str, capitole: list[dict], nota: str = "") -> list[dict]:
    stubs: list[dict] = []
    for cap in capitole:
        subs = cap.get("subcapitole", [])
        if subs:
            for sub in subs:
                stubs.append(_stub(modul, cap["titlu"], sub["titlu"]))
        else:
            stubs.append(_stub(modul, cap["titlu"], ""))
    if nota:
        for s in stubs:
            s["observatii"] = nota
    return stubs


async def _generate_chunk_ai(modul: str, capitole: list[dict], part: int, total: int) -> list[dict]:
    content = await llm_client.chat(
        _SYSTEM_PROMPT, _chunk_prompt(modul, capitole, part, total), max_tokens=OUT_TOKENS_PER_CALL
    )
    data = llm_client.parse_json(content)
    items = data.get("scenarii", [])
    if not items:
        raise ValueError("Răspuns AI fără scenarii")
    # Mistral trimite null pentru campurile goale; Pydantic aplica default-urile
    # doar pentru chei absente, deci eliminam valorile None inainte de validare.
    return [
        _Scenariu(**{k: v for k, v in s.items() if v is not None}).model_dump()
        for s in items
    ]


def _structure_chars(structure: dict[str, list[dict]]) -> int:
    total = 0
    for modul, capitole in structure.items():
        total += len(modul)
        for cap in capitole:
            total += len(cap["titlu"]) + sum(len(t) for t in cap["text"])
            for sub in cap["subcapitole"]:
                total += len(sub["titlu"]) + sum(len(t) for t in sub["text"])
    return total


def estimate_scenarii_job(docx_path: Path) -> dict:
    """Pre-check: tokeni estimați, apeluri, module și dacă încape în bugetul zilnic."""
    structure = _extract_structure(docx_path)
    chunks = _pack_chunks(structure)
    calls = max(1, len(chunks))
    input_tokens = llm_client.estimate_tokens(" " * _structure_chars(structure))
    est_tokens = input_tokens + calls * (CALL_OVERHEAD_TOKENS + OUT_TOKENS_PER_CALL)
    return {
        "est_tokens": est_tokens,
        "calls": calls,
        "modules": max(1, len(structure)),
        "est_minutes": max(1, round(calls * 30 / 60)),
        "fits_budget": est_tokens <= llm_client.remaining_budget(),
    }


def _stub(modul: str, capitol: str, subcapitol: str) -> dict:
    titlu = subcapitol or capitol
    return {
        "capitol": capitol,
        "subcapitol": subcapitol,
        "titlu_scenariu": f"Verificare: {titlu}",
        "obiectiv": f"Verifică funcționalitatea '{titlu}' din modulul {modul}.",
        "preconditii": "• Utilizator autentificat cu drepturi pe modul\n• Date de test pregătite",
        "pasi": "1. Navigare la funcționalitate\n2. Execuție flux conform specificației\n3. Validare rezultat",
        "rezultat_asteptat": "Sistemul procesează fluxul conform specificației. Nu apar erori.",
        "tip_test": "Funcțional - Pozitiv",
        "prioritate": "High" if subcapitol else "Critical",
        "dependente": "—",
        "observatii": "",
    }


def _write_excel(scenarios: list[dict]) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scenarii"

    headers = [
        "ID", "Capitol", "Subcapitol", "Titlu Scenariu",
        "Obiectiv", "Precondiții", "Pași de Execuție", "Rezultat Așteptat",
        "Tip Test", "Prioritate", "Dependente", "Observații",
    ]
    header_fill = PatternFill(fill_type="solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wrap = Alignment(wrap_text=True, vertical="top")
    for i, s in enumerate(scenarios, start=1):
        ws.append([
            s["id"],
            s["capitol"],
            s["subcapitol"],
            s["titlu_scenariu"],
            s["obiectiv"],
            s["preconditii"],
            s["pasi"],
            s["rezultat_asteptat"],
            s["tip_test"],
            s["prioritate"],
            s["dependente"],
            s["observatii"],
        ])
        for col_idx in range(1, 13):
            ws.cell(row=i + 1, column=col_idx).alignment = wrap

    col_widths = [10, 25, 25, 35, 35, 30, 40, 35, 20, 12, 15, 20]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    ws.freeze_panes = "B2"

    fd, tmp_name = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    output_path = Path(tmp_name)
    wb.save(str(output_path))
    return output_path


async def run_scenarii_pipeline(
    docx_path: Path,
    use_ai: bool = False,
    on_step=None,
) -> tuple[Path, list[dict]]:
    """DOCX spec → Excel cu scenarii de testare. Returns (xlsx_path, scenarios).

    use_ai=True: un apel Mistral per bucată (chunk), cu fallback la stub-uri per bucată.
    use_ai=False: stub-urile deterministe de azi (plan B garantat).
    """
    structure = _extract_structure(docx_path)

    scenarios: list[dict] = []
    if use_ai:
        chunks = _pack_chunks(structure)
        # numarul partii in cadrul modulului, pentru antetul promptului
        per_module_totals: dict[str, int] = {}
        for ch in chunks:
            per_module_totals[ch["modul"]] = per_module_totals.get(ch["modul"], 0) + 1
        per_module_seen: dict[str, int] = {}

        for idx, ch in enumerate(chunks, start=1):
            modul = ch["modul"]
            per_module_seen[modul] = per_module_seen.get(modul, 0) + 1
            if on_step:
                on_step(f"chunk:{idx}/{len(chunks)}:{modul}")
            try:
                generated = await _generate_chunk_ai(
                    modul, ch["capitole"],
                    part=per_module_seen[modul], total=per_module_totals[modul],
                )
                for s in generated:
                    s["ai"] = True
                scenarios.extend(generated)
            except Exception:
                fallback = _module_stubs(
                    modul, ch["capitole"],
                    nota="Generat fără AI (fallback — apelul AI a eșuat)",
                )
                for s in fallback:
                    s["ai"] = False
                scenarios.extend(fallback)
    else:
        for modul, capitole in structure.items():
            stubs = _module_stubs(modul, capitole)
            for s in stubs:
                s["ai"] = False
            scenarios.extend(stubs)

    if not scenarios:
        empty = {
            "capitol": "General",
            "subcapitol": "",
            "titlu_scenariu": "Verificare generală",
            "obiectiv": "Verifică funcționalitățile generale din specificație.",
            "preconditii": "• Utilizator autentificat",
            "pasi": "1. Navigare la funcționalitate\n2. Execuție flux\n3. Validare",
            "rezultat_asteptat": "Funcționare conform specificației.",
            "tip_test": "Funcțional - Pozitiv",
            "prioritate": "High",
            "dependente": "—",
            "observatii": "Nicio structură de headings detectată în document.",
            "ai": False,
        }
        scenarios.append(empty)

    for i, s in enumerate(scenarios, start=1):
        s["id"] = f"TC-{i:03d}"

    if on_step:
        on_step("building")
    return _write_excel(scenarios), scenarios
