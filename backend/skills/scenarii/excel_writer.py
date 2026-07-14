# -*- coding: utf-8 -*-
"""Generatorul Excel pentru scenariile de testare — formatul validat (v3):
- sheet per modul cu dropdown Status (Netestat/Testat cu succes/Testat cu eroare)
  si colorare conditionata (galben/verde/rosu)
- avertizare portocalie: scenariu "Testat cu succes" cu dependente netestate
- scenariile din cerintele specifice ale clientului: randuri galbene
- Overview cu contoare live (COUNTIF)
- 🗺️ Plan Executie (ordine topologica cross-module)
- 🚫 Excluse vs Standard (capitole standard absente din spec-ul clientului)
"""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

COLUMNS = [
    "ID", "Modul", "Capitol", "Subcapitol", "Scenariu", "Obiectiv",
    "Pasi de Test", "Rezultat Asteptat", "Tip Test", "Prioritate",
    "Dependente", "Status", "Observatii",
]
STATUS_COL = "L"
STATUS_VALUES = ["Netestat", "Testat cu succes", "Testat cu eroare"]
WIDTHS = [12, 26, 30, 34, 34, 34, 42, 36, 18, 11, 42, 16, 26]

HFILL = PatternFill(fill_type="solid", fgColor="1F3864")
HFONT = Font(bold=True, color="FFFFFF", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
YELLOW = PatternFill(fill_type="solid", fgColor="FFF2A8")
GREEN = PatternFill(fill_type="solid", fgColor="C6EFCE")
RED = PatternFill(fill_type="solid", fgColor="FFC7CE")
ORANGE = PatternFill(fill_type="solid", fgColor="FFD8A8")
SPECIFIC_FILL = PatternFill(fill_type="solid", fgColor="FFF9C4")


def _short(t: str, n: int = 45) -> str:
    return t if len(t) <= n else t[: n - 1].rstrip() + "…"


def _topo_order(catalog: dict, id_info: dict) -> list[str]:
    all_ids = [s["ID"] for items in catalog.values() for s in items]
    pos = {sid: i for i, sid in enumerate(all_ids)}
    visited: dict[str, int] = {}
    order: list[str] = []

    def visit(sid: str):
        state = visited.get(sid)
        if state is not None:
            return
        visited[sid] = 0
        for dep in id_info.get(sid, {}).get("deps", []):
            if dep in pos and visited.get(dep) != 0:
                visit(dep)
        visited[sid] = 1
        order.append(sid)

    for sid in sorted(all_ids, key=lambda x: pos[x]):
        visit(sid)
    return order


def write_scenarii_excel(
    catalog: dict,
    output_path: Path,
    excluded: list[tuple[str, dict]] | None = None,
    client_name: str = "",
) -> None:
    """catalog: {sheet: [scenariu]} — scenariile au deps (list), _specific (bool)."""
    excluded = excluded or []

    # index ID -> sheet/rand/titlu (dupa excludere, randurile sunt finale)
    id_info: dict[str, dict] = {}
    for sheet, items in catalog.items():
        for idx, s in enumerate(items):
            id_info[s["ID"]] = {
                "sheet": sheet, "row": idx + 2,
                "title": s.get("Scenariu", ""), "deps": s.get("deps", []),
            }

    # afisare Dependente: "ID (titlu scurt)" pe randuri
    for items in catalog.values():
        for s in items:
            deps = s.get("deps", [])
            s["Dependente"] = "\n".join(
                f"{d} ({_short(id_info[d]['title'])})" if d in id_info else d
                for d in deps
            ) if deps else "—"

    wb = openpyxl.Workbook()
    ov = wb.active
    ov.title = "Overview"

    sheet_rows = []
    for sheet, items in catalog.items():
        if not items:
            continue
        ws = wb.create_sheet(sheet)
        ws.append(COLUMNS)
        for cell in ws[1]:
            cell.fill = HFILL
            cell.font = HFONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for i, s in enumerate(items, start=2):
            ws.append([s.get(c, "") for c in COLUMNS])
            for col in range(1, len(COLUMNS) + 1):
                cell = ws.cell(row=i, column=col)
                cell.alignment = WRAP
                if s.get("_specific"):
                    cell.fill = SPECIFIC_FILL
        for idx, w in enumerate(WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = w
        ws.freeze_panes = "E2"

        last = len(items) + 1
        rng = f"{STATUS_COL}2:{STATUS_COL}{last}"
        dv = DataValidation(
            type="list", formula1='"' + ",".join(STATUS_VALUES) + '"',
            allow_blank=True, showDropDown=False,
        )
        ws.add_data_validation(dv)
        dv.add(rng)

        # avertizare: succes cu dependente ne-testate → rand portocaliu
        for idx, s in enumerate(items):
            deps = [d for d in s.get("deps", []) if d in id_info]
            if not deps:
                continue
            r = idx + 2
            conds = [
                f"'{id_info[d]['sheet']}'!${STATUS_COL}${id_info[d]['row']}<>\"Testat cu succes\""
                for d in deps
            ]
            formula = f'=AND(${STATUS_COL}{r}="Testat cu succes",OR({",".join(conds)}))'
            ws.conditional_formatting.add(
                f"A{r}:M{r}", FormulaRule(formula=[formula], fill=ORANGE, stopIfTrue=True))

        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=['"Netestat"'], fill=YELLOW))
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=['"Testat cu succes"'], fill=GREEN))
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=['"Testat cu eroare"'], fill=RED))
        sheet_rows.append((sheet, items, last))

    # ── Overview ──
    title = "SCENARII DE TESTARE - CHARISMA ERP"
    if client_name:
        title += f" — {client_name}"
    ov.append(["", title, "", "", "", ""])
    ov["B1"].font = Font(bold=True, size=14, color="1F3864")
    ov.append(["", "Catalog standard + cerinte specifice client (randuri galbene). Contoarele se actualizeaza automat.", "", "", "", ""])
    ov.append([])
    ov.append(["", "Modul", "Nr. Scenarii", "Netestat", "Testat cu succes", "Testat cu eroare"])
    for c in ov[4]:
        if c.value:
            c.fill = HFILL
            c.font = HFONT
    r = 5
    for sheet, items, last in sheet_rows:
        rng = f"'{sheet}'!${STATUS_COL}$2:${STATUS_COL}${last}"
        spec_count = sum(1 for s in items if s.get("_specific"))
        label = f"{sheet}  (+{spec_count} specifice)" if spec_count else sheet
        ov.append([
            "", label, len(items),
            f'=COUNTIF({rng},"Netestat")',
            f'=COUNTIF({rng},"Testat cu succes")',
            f'=COUNTIF({rng},"Testat cu eroare")',
        ])
        ov.cell(row=r, column=4).fill = YELLOW
        ov.cell(row=r, column=5).fill = GREEN
        ov.cell(row=r, column=6).fill = RED
        r += 1
    for idx, w in enumerate([3, 34, 13, 12, 16, 16], start=1):
        ov.column_dimensions[get_column_letter(idx)].width = w

    # ── Plan Executie ──
    plan = wb.create_sheet("🗺️ Plan Executie")
    plan.append(["Ordine", "ID", "Modul", "Scenariu", "Dependente", "Capitol"])
    for c in plan[1]:
        c.fill = HFILL
        c.font = HFONT
    all_scen = {s["ID"]: s for items in catalog.values() for s in items}
    for i, sid in enumerate(_topo_order(catalog, id_info), start=1):
        s = all_scen[sid]
        plan.append([
            i, sid, id_info[sid]["sheet"], s.get("Scenariu", ""),
            ", ".join(s.get("deps", [])) or "—", s.get("Capitol", ""),
        ])
        if s.get("_specific"):
            for c in plan[plan.max_row]:
                c.fill = SPECIFIC_FILL
    for idx, w in enumerate([9, 12, 24, 46, 24, 30], start=1):
        plan.column_dimensions[get_column_letter(idx)].width = w
    plan.freeze_panes = "A2"

    # ── Excluse vs Standard ──
    ex = wb.create_sheet("🚫 Excluse vs Standard")
    ex.append(["ID", "Modul", "Scenariu", "Capitol standard absent din specificatia clientului"])
    for c in ex[1]:
        c.fill = HFILL
        c.font = HFONT
    if excluded:
        for sheet, s in excluded:
            ex.append([s["ID"], sheet, s.get("Scenariu", ""), s.get("std_ref", "")])
            for c in ex[ex.max_row]:
                c.fill = RED
                c.alignment = WRAP
    else:
        ex.append(["—", "—", "Nimic exclus: specificatia acopera toate capitolele standard.", "—"])
    for idx, w in enumerate([12, 24, 46, 62], start=1):
        ex.column_dimensions[get_column_letter(idx)].width = w

    wb.save(str(output_path))
