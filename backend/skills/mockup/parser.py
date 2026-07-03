from __future__ import annotations
from dataclasses import dataclass, field
from pathlib import Path
import zipfile
import defusedxml.ElementTree as ET
import openpyxl


@dataclass
class FilterField:
    label: str
    mandatory: bool = False
    coord: str = ""


@dataclass
class ButtonDef:
    label: str
    group: str  # "direct" | "actiuni"
    coord: str = ""


@dataclass
class ColumnDef:
    name: str
    coord: str = ""


@dataclass
class Section:
    title: str
    filter_fields: list[FilterField] = field(default_factory=list)
    buttons: list[ButtonDef] = field(default_factory=list)
    columns: list[ColumnDef] = field(default_factory=list)
    data_rows: list[list] = field(default_factory=list)
    filter_values: dict[str, str] = field(default_factory=dict)  # label → example value from col B


@dataclass
class ScreenSpec:
    screen_title: str
    source_file: str
    sections: list[Section] = field(default_factory=list)
    observatii: list[str] = field(default_factory=list)
    reference_image: bytes | None = None
    cell_comments: dict[str, str] = field(default_factory=dict)


from .config import MANDATORY_FILTERS


def _is_date_or_number(val) -> bool:
    import datetime
    return isinstance(val, (int, float, datetime.datetime, datetime.date))


def _is_action_cell(val) -> bool:
    return isinstance(val, str) and val.strip().startswith("Actiuni:")


def _parse_action_buttons(val: str) -> list[ButtonDef]:
    lines = val.strip().splitlines()
    return [ButtonDef(label=line.strip(), group="actiuni")
            for line in lines[1:] if line.strip()]


def _row_values(ws, row_idx: int) -> list:
    return [ws.cell(row=row_idx, column=c).value
            for c in range(1, ws.max_column + 1)]


def _col_letter(col_idx: int) -> str:
    """Convert 1-based column index to Excel letter (1→A, 26→Z, 27→AA)."""
    result = ""
    while col_idx > 0:
        col_idx, rem = divmod(col_idx - 1, 26)
        result = chr(65 + rem) + result
    return result


def _coord(row_idx: int, col_idx: int) -> str:
    return f"{_col_letter(col_idx)}{row_idx}"


_DIRECT_BTNS = {"Filtreaza", "Adauga", "Importa", "Exporta", "Selecteaza",
                "Reseteaza", "Salveaza layout", "Inchide"}
_SECTION_MARKERS = {"Zona de filtrare", "Filtre"}


def _fmt_val(val) -> str:
    """Format a cell value for display in mockup."""
    import datetime
    if val is None:
        return ""
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%d.%m.%Y")
    return str(val)


def _parse_ecran(ws) -> tuple[str, list[Section]]:
    sections: list[Section] = []
    current: Section | None = None
    screen_title = ""
    filter_labels_seen: list[str] = []
    grid_col_start: int = 0
    grid_header_row_idx: int = 0
    last_data_row_idx: int = 0
    data_collection_frozen: bool = False

    for row_idx in range(1, ws.max_row + 1):
        row = _row_values(ws, row_idx)
        non_none = [v for v in row if v is not None]
        if not non_none:
            # Empty row after data rows → freeze main section data collection
            if last_data_row_idx > 0:
                data_collection_frozen = True
            continue

        col_a = row[0]

        # New section: "Zona de filtrare" or "Filtre"
        if isinstance(col_a, str) and col_a.strip() in _SECTION_MARKERS:
            if current is not None:
                sections.append(current)
            # Title is in col C (index 2)
            title = str(row[2]).strip() if len(row) > 2 and row[2] else (
                str(row[1]).strip() if len(row) > 1 and row[1] else ""
            )
            screen_title = screen_title or title
            current = Section(title=title)
            filter_labels_seen = []
            grid_col_start = 0
            grid_header_row_idx = 0
            last_data_row_idx = 0
            data_collection_frozen = False
            continue

        if current is None:
            continue

        # Col A: filter field names (short strings, not dates/numbers)
        if isinstance(col_a, str) and not _is_date_or_number(col_a):
            label = col_a.strip()
            if label and label not in filter_labels_seen:
                filter_labels_seen.append(label)
                # Bold font → mandatory; fallback to hardcoded MANDATORY_FILTERS
                cell_bold = ws.cell(row=row_idx, column=1).font.bold
                is_mandatory = bool(cell_bold) or (label in MANDATORY_FILTERS)
                current.filter_fields.append(
                    FilterField(
                        label=label,
                        mandatory=is_mandatory,
                        coord=_coord(row_idx, 1),
                    )
                )
                # Store example value from col B if present
                col_b_val = row[1] if len(row) > 1 else None
                if col_b_val is not None:
                    current.filter_values[label] = _fmt_val(col_b_val)

        # Scan remaining cells for buttons
        for col_idx, cell_val in enumerate(row[1:], start=2):
            if cell_val is None:
                continue
            if _is_action_cell(cell_val):
                btns = _parse_action_buttons(str(cell_val))
                for b in btns:
                    b.coord = _coord(row_idx, col_idx)
                current.buttons.extend(btns)
                continue
            if isinstance(cell_val, str) and not _is_date_or_number(cell_val):
                label = cell_val.strip()
                if label in _DIRECT_BTNS:
                    if not any(b.label == label for b in current.buttons):
                        current.buttons.append(
                            ButtonDef(label=label, group="direct", coord=_coord(row_idx, col_idx))
                        )

        # Grid header detection: 3+ string cells from col 3 onwards, no button labels
        str_from_c = [
            (col_idx + 3, v)
            for col_idx, v in enumerate(row[2:])
            if isinstance(v, str) and not _is_date_or_number(v) and v.strip()
            and v.strip() not in _DIRECT_BTNS and not _is_action_cell(v)
        ]
        if len(str_from_c) >= 3 and not current.columns:
            grid_col_start = str_from_c[0][0]
            grid_header_row_idx = row_idx
            for abs_col, v in str_from_c:
                current.columns.append(
                    ColumnDef(name=v.strip(), coord=_coord(row_idx, abs_col))
                )
            continue  # skip data check on the header row itself

        # Data row: any row after grid header, with values, before the first empty-row gap
        if (grid_col_start > 0 and grid_header_row_idx > 0
                and row_idx > grid_header_row_idx
                and current.columns
                and not data_collection_frozen):
            data = [
                _fmt_val(row[grid_col_start - 1 + i]) if grid_col_start - 1 + i < len(row) else ""
                for i in range(len(current.columns))
            ]
            if any(d for d in data):
                current.data_rows.append(data)
                last_data_row_idx = row_idx

    # Finalize main section BEFORE secondary sections (preserves order)
    if current is not None and not any(s.title == current.title for s in sections):
        sections.append(current)

    # Secondary sections: isolated single long-text cell followed by column header row
    for row_idx in range(1, ws.max_row + 1):
        row = _row_values(ws, row_idx)
        non_none = [v for v in row if v is not None]
        # Single non-empty cell anywhere in the row (check col C area too)
        str_vals = [v for v in non_none if isinstance(v, str)]
        if len(non_none) == 1 and str_vals and len(str_vals[0]) > 15:
            label = str_vals[0].strip()
            if label not in _SECTION_MARKERS and not any(s.title == label for s in sections):
                if row_idx < ws.max_row:
                    next_row = _row_values(ws, row_idx + 1)
                    next_str = [(col_idx + 1, v) for col_idx, v in enumerate(next_row)
                                if isinstance(v, str) and v.strip()]
                    if len(next_str) >= 2:
                        sec = Section(title=label)
                        sec_col_start = next_str[0][0]
                        for abs_col, v in next_str:
                            sec.columns.append(ColumnDef(name=v.strip(), coord=_coord(row_idx + 1, abs_col)))
                        # Collect data rows for secondary section
                        for data_row_idx in range(row_idx + 2, ws.max_row + 1):
                            data_row = _row_values(ws, data_row_idx)
                            data_non_none = [v for v in data_row if v is not None]
                            if not data_non_none:
                                break
                            data = [_fmt_val(data_row[sec_col_start - 1 + i]) if sec_col_start - 1 + i < len(data_row) else ""
                                    for i in range(len(sec.columns))]
                            if any(d for d in data):
                                sec.data_rows.append(data)
                        sections.append(sec)

    return screen_title, sections


def _parse_observatii(ws) -> list[str]:
    lines = []
    for row in ws.iter_rows(values_only=True):
        val = row[0] if row else None
        if val is not None:
            lines.append(str(val).strip())
    return [line for line in lines if line]


def _extract_reference_image(ws) -> bytes | None:
    if not getattr(ws, "_images", None):
        return None
    try:
        return ws._images[0]._data()
    except Exception:
        return None


def _read_threaded_comments(xlsx_path) -> dict[str, str]:
    comments: dict[str, str] = {}
    try:
        with zipfile.ZipFile(str(xlsx_path)) as z:
            for name in z.namelist():
                if "threadedComment" in name and name.endswith(".xml"):
                    with z.open(name) as f:
                        content = f.read().decode("utf-8")
                    root = ET.fromstring(content)
                    ns = {"tc": "http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments"}
                    for tc in root.findall("tc:threadedComment", ns):
                        ref = tc.get("ref")
                        text_el = tc.find("tc:text", ns)
                        if ref and text_el is not None and text_el.text:
                            comments[ref] = text_el.text.strip()
    except Exception:
        pass
    return comments


def parse_excel(path) -> ScreenSpec:
    wb = openpyxl.load_workbook(str(path), data_only=True)
    sheets = set(wb.sheetnames)

    screen_title, sections = _parse_ecran(wb["Ecran"])
    observatii = _parse_observatii(wb["Observatii"]) if "Observatii" in sheets else []
    ref_image = _extract_reference_image(wb["exemplu ecran grafic"]) if "exemplu ecran grafic" in sheets else None
    cell_comments = _read_threaded_comments(path)

    return ScreenSpec(
        screen_title=screen_title,
        source_file=Path(path).name,
        sections=sections,
        observatii=observatii,
        reference_image=ref_image,
        cell_comments=cell_comments,
    )
